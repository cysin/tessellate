#!/usr/bin/env python3
"""
Test GDRR (Goal-Driven Ruin and Recreate) algorithm on manual1.xlsx.

GDRR is inspired by: https://github.com/JeroenGar/gdrr-2bp.git
Paper: "A goal-driven ruin and recreate heuristic for the 2D variable-sized
        bin packing problem with guillotine constraints"
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Problem, Item, Bin
from tessellate.algorithms.gdrr_packer import GDRRPacker, GDRRLightPacker
from tessellate.algorithms.skyline import SkylinePacker
from tessellate.algorithms.local_search_packer import LocalSearchPacker


def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            normalized_headers[column_mapping[header]] = idx

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        grain_mapping = {'mixed': True, 'fixed': False, '可旋转': True, '不可旋转': False}

        item = Item(
            id=row[normalized_headers['Code']],
            width=int(float(row[normalized_headers['Width']])),
            height=int(float(row[normalized_headers['Height']])),
            thickness=int(float(row[normalized_headers['Thickness']])),
            material=row[normalized_headers['Color']],
            quantity=int(float(row[normalized_headers['Qty']])),
            rotatable=grain_mapping.get(row[normalized_headers['Grain']], True)
        )
        items.append(item)

    return items


def main():
    excel_path = Path("test_data/bench/manual1.xlsx")

    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        return

    print("=" * 70)
    print("Testing GDRR Algorithm on manual1.xlsx")
    print("=" * 70)
    print()

    items = load_excel_manual1(excel_path)

    print(f"Loaded {len(items)} item types")
    print(f"Total items: {sum(item.quantity for item in items)}")
    print()

    # Create bins matching item specifications
    thicknesses = list(set(item.thickness for item in items))
    materials = list(set(item.material for item in items))

    bins = []
    for thickness in thicknesses:
        for material in materials:
            bins.append(
                Bin(
                    id=f"Board-2440x1220-{thickness}mm-{material}",
                    width=2440,
                    height=1220,
                    thickness=thickness,
                    material=material
                )
            )

    problem = Problem(items=items, bins=bins, kerf=3.0)

    # Test algorithms
    print("=" * 70)
    print("TESTING ALGORITHMS")
    print("=" * 70)
    print()

    algorithms = [
        ("Skyline (baseline)", SkylinePacker(time_limit=5.0, use_min_waste=True)),
        ("Local Search (old)", LocalSearchPacker(time_limit=60.0, max_iterations=300)),
        ("GDRR-Light (30s)", GDRRLightPacker(time_limit=30.0)),
        ("GDRR (60s)", GDRRPacker(time_limit=60.0, iterations=500, history_length=100)),
    ]

    results = []

    for name, algorithm in algorithms:
        print(f"\n{'='*70}")
        print(f"Testing: {name}")
        print(f"{'='*70}")

        solution = algorithm.solve(problem)

        num_bins = solution.num_bins()
        utilization = solution.total_utilization()
        items_placed = sum(len(bp.items) for bp in solution.bins)
        items_unplaced = sum(qty for _, qty in solution.unplaced)
        exec_time = solution.metadata.get('execution_time', 0)

        results.append({
            'name': name,
            'bins': num_bins,
            'utilization': utilization,
            'placed': items_placed,
            'unplaced': items_unplaced,
            'time': exec_time,
            'iterations': solution.metadata.get('iterations', '-'),
            'improved': solution.metadata.get('improved', '-')
        })

        print()

    # Summary
    print()
    print("=" * 70)
    print("RESULTS SUMMARY")
    print("=" * 70)
    print()
    print(f"{'Algorithm':<25} {'Bins':>6} {'Util%':>7} {'Time':>7} {'Iter':>6} {'Impr':>6}")
    print("-" * 70)

    for result in results:
        iter_str = str(result['iterations']) if result['iterations'] != '-' else '-'
        impr_str = str(result['improved']) if result['improved'] != '-' else '-'
        print(f"{result['name']:<25} {result['bins']:>6} {result['utilization']:>6.2%} "
              f"{result['time']:>6.1f}s {iter_str:>6} {impr_str:>6}")

    print()
    print("Target: 10 boards (manual solution)")
    print()

    # Find best result
    best = min(results, key=lambda x: (x['bins'], -x['utilization']))
    print(f"Best result: {best['name']}")
    print(f"  {best['bins']} boards @ {best['utilization']:.2%} utilization")

    if best['bins'] == 10:
        print()
        print("🎉 SUCCESS: Matched manual solution target!")
    elif best['bins'] == 11:
        print()
        print("Result: 11 boards (+1 vs manual target)")
        print("This is typical for greedy/metaheuristic algorithms")
        print()
        print("GDRR provides exploration beyond greedy but may not improve")
        print("on well-optimized initial solutions for this particular dataset.")


if __name__ == "__main__":
    main()
