#!/usr/bin/env python3
"""
Detailed analysis of manual1.jpg solution to understand the exact packing strategy.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Problem, Item, Bin


def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            normalized_headers[column_mapping[header]] = idx

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        grain_mapping = {'mixed': True, 'fixed': False, '可旋转': True, '不可旋转': False}

        item = Item(
            id=row[normalized_headers['Code']],
            width=int(float(row[normalized_headers['Width']])),
            height=int(float(row[normalized_headers['Height']])),
            thickness=int(float(row[normalized_headers['Thickness']])),
            material=row[normalized_headers['Color']],
            quantity=int(float(row[normalized_headers['Qty']])),
            rotatable=grain_mapping.get(row[normalized_headers['Grain']], True)
        )
        items.append(item)

    return items


def analyze_strip_packing():
    """Analyze how items could be packed into strips optimally."""

    excel_path = Path("test_data/bench/manual1.xlsx")
    items = load_excel_manual1(excel_path)

    print("="*70)
    print("DETAILED MANUAL SOLUTION ANALYSIS")
    print("="*70)
    print()

    # Extract item details
    item_details = []
    for item in items:
        for _ in range(item.quantity):
            item_details.append({
                'width': item.width,
                'height': item.height,
                'id': item.id
            })

    # Constants
    BIN_WIDTH = 2440
    BIN_HEIGHT = 1220
    STRIP_HEIGHT = 554
    KERF = 3
    STRIPS_PER_BIN = 2

    print(f"Items: {len(item_details)} total")
    print(f"Bin: {BIN_WIDTH}mm × {BIN_HEIGHT}mm")
    print(f"Strip height: {STRIP_HEIGHT}mm")
    print(f"Strips per bin: {STRIPS_PER_BIN}")
    print(f"Kerf: {KERF}mm")
    print()

    # Count items by width
    width_counts = {}
    for item in item_details:
        width = item['width']
        width_counts[width] = width_counts.get(width, 0) + 1

    print("Item distribution:")
    for width in sorted(width_counts.keys(), reverse=True):
        count = width_counts[width]
        print(f"  {width}mm: {count} items")
    print()

    # Analyze theoretical minimum strips using FFD
    print("="*70)
    print("THEORETICAL ANALYSIS")
    print("="*70)
    print()

    # Sort items by width descending
    sorted_items = sorted(item_details, key=lambda x: x['width'], reverse=True)

    # Run FFD strip packing
    strips = []
    for item in sorted_items:
        width = item['width']

        # Find first strip that fits
        placed = False
        for strip in strips:
            space_available = BIN_WIDTH - strip['used']
            needed = width if strip['used'] == 0 else width + KERF

            if needed <= space_available:
                strip['items'].append(item)
                strip['used'] += needed
                placed = True
                break

        if not placed:
            strips.append({'items': [item], 'used': width})

    print(f"FFD strip packing: {len(strips)} strips")
    print(f"Bins needed (2 strips/bin): {(len(strips) + 1) // 2} bins")
    print()

    # Analyze strip utilization
    print("Strip utilization:")
    total_waste = 0
    for i, strip in enumerate(strips):
        waste = BIN_WIDTH - strip['used']
        util = strip['used'] / BIN_WIDTH * 100
        total_waste += waste
        if i < 10 or waste > 200:  # Show first 10 and any with high waste
            widths = [item['width'] for item in strip['items']]
            print(f"  Strip {i+1:2d}: {strip['used']:4d}mm used, {waste:4d}mm waste ({util:.1f}%) - widths: {widths}")

    print()
    print(f"Total waste across all strips: {total_waste}mm")
    print(f"Average waste per strip: {total_waste / len(strips):.1f}mm")
    print()

    # Check if we can improve by reordering
    print("="*70)
    print("OPTIMIZATION OPPORTUNITY ANALYSIS")
    print("="*70)
    print()

    # Find strips with high waste that might be combinable
    high_waste_strips = [(i, s) for i, s in enumerate(strips) if BIN_WIDTH - s['used'] > 300]
    print(f"Strips with >300mm waste: {len(high_waste_strips)}")
    for i, strip in high_waste_strips[:5]:
        waste = BIN_WIDTH - strip['used']
        widths = [item['width'] for item in strip['items']]
        print(f"  Strip {i+1}: {waste}mm waste - {widths}")
    print()

    # Calculate theoretical minimum based on total area
    total_item_area = sum(item['width'] * item['height'] for item in item_details)
    total_kerf_area_per_strip = (len(item_details) / len(strips)) * KERF * STRIP_HEIGHT  # rough estimate
    bin_area = BIN_WIDTH * BIN_HEIGHT

    print(f"Total item area: {total_item_area:,} mm²")
    print(f"Bin area: {bin_area:,} mm²")
    print(f"Theoretical minimum bins (by area): {total_item_area / bin_area:.2f}")
    print(f"Theoretical minimum (accounting for strip constraint): {len(strips) / STRIPS_PER_BIN:.1f} = {(len(strips) + 1) // 2} bins")
    print()

    print("="*70)
    print("CONCLUSION")
    print("="*70)
    print()
    print(f"FFD achieves: {(len(strips) + 1) // 2} bins")
    print(f"Manual solution: 10 bins (20 strips)")
    print(f"Gap: {(len(strips) + 1) // 2 - 10} bins ({len(strips) - 20} strips)")
    print()

    if len(strips) > 20:
        print(f"To match manual solution, we need to reduce from {len(strips)} to 20 strips.")
        print(f"This requires better packing optimization to reduce waste.")
        print()
        print("Possible strategies:")
        print("1. Better item ordering/grouping")
        print("2. Bin Completion algorithms (try to fill bins optimally)")
        print("3. Integer Linear Programming")
        print("4. Metaheuristic optimization (genetic algorithms, simulated annealing)")


if __name__ == "__main__":
    analyze_strip_packing()
