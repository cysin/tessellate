#!/usr/bin/env python3
"""
Analyze the manual1.jpg solution pattern to derive the algorithm.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Item

def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            normalized_headers[column_mapping[header]] = idx

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        grain_mapping = {'mixed': True, 'fixed': False, '可旋转': True, '不可旋转': False}

        item = Item(
            id=row[normalized_headers['Code']],
            width=int(float(row[normalized_headers['Width']])),
            height=int(float(row[normalized_headers['Height']])),
            thickness=int(float(row[normalized_headers['Thickness']])),
            material=row[normalized_headers['Color']],
            quantity=int(float(row[normalized_headers['Qty']])),
            rotatable=grain_mapping.get(row[normalized_headers['Grain']], True)
        )
        items.append(item)

    return items


def main():
    print("=" * 70)
    print("MANUAL SOLUTION PATTERN ANALYSIS")
    print("=" * 70)
    print()

    items = load_excel_manual1("test_data/bench/manual1.xlsx")

    print("OBSERVATION 1: All items have same HEIGHT")
    print("-" * 70)
    for item in items:
        print(f"{item.id}: {item.width}x{item.height}mm × {item.quantity}")
    print()
    print("Notice: ALL items are 554mm tall!")
    print("Manual solution ROTATES all items so 554mm becomes the HEIGHT")
    print()

    print("OBSERVATION 2: Items vary by WIDTH only")
    print("-" * 70)
    widths = [(item.width, item.quantity) for item in items]
    widths.sort(reverse=True)
    print("Sorted by width (descending):")
    for width, qty in widths:
        print(f"  {width}mm × {qty} items")
    print()

    print("OBSERVATION 3: Bin dimensions")
    print("-" * 70)
    print("Bin: 2440mm (width) × 1220mm (height)")
    print("With 554mm item height, we can fit: 1220 / 554 = 2.2 strips")
    print("So maximum 2 strips per bin (with kerf)")
    print()

    print("OBSERVATION 4: Manual solution strategy")
    print("-" * 70)
    print("1. Rotate ALL items to landscape: 554mm height (uniform)")
    print("2. Items widths: 336-864mm")
    print("3. Pack items HORIZONTALLY in strips")
    print("4. 2 horizontal strips per board (554mm × 2 ≈ 1220mm)")
    print("5. Each strip width ≈ 2440mm")
    print()

    print("OBSERVATION 5: Packing pattern per strip")
    print("-" * 70)
    print("Looking at manual1.jpg:")
    print()
    print("Board 1 (top strip):")
    print("  864 + 832 + 800 + 768 + 736 = 4000mm (too wide!)")
    print("  Actual: Items arranged with VERTICAL stacking on LEFT")
    print("          Plus 736mm items on RIGHT in separate column")
    print()
    print("Board 1 (bottom strip):")
    print("  336 item")
    print()

    print("KEY INSIGHT: TWO-COLUMN LAYOUT")
    print("-" * 70)
    print("Manual solution uses:")
    print("  LEFT column: Variable width (~1700mm)")
    print("  RIGHT column: Fixed for 736mm items (~740mm)")
    print()
    print("This creates a STRUCTURED GUILLOTINE pattern!")
    print()

    print("DERIVED ALGORITHM: 'Manual Pattern' Packer")
    print("=" * 70)
    print("1. Force rotation: ALL items to 554mm height (landscape)")
    print("2. Sort items by width DESCENDING")
    print("3. Identify most common item (736mm, qty=20)")
    print("4. Use TWO-COLUMN layout:")
    print("   - RIGHT column: Reserve for 736mm items")
    print("   - LEFT column: Pack remaining items by width")
    print("5. Pack in HORIZONTAL STRIPS (2 per board)")
    print("6. Within each strip, pack items LEFT to RIGHT")
    print()

    print("Expected result: 10 boards (matching manual)")
    print()


if __name__ == "__main__":
    main()
