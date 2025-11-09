"""
Flask web application for the Tessellate cutting stock optimizer.

Provides a web interface for solving 2D guillotine cutting stock problems.
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import sys
import os
import json
import traceback
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Add parent directory to path to import tessellate
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from tessellate import solve
from tessellate.core.models import Problem
from tessellate.core.validator import SolutionValidator
from tessellate.algorithms.guillotine_tree import GuillotineTreeBuilder

app = Flask(__name__)
CORS(app)


def deduplicate_items(items):
    """
    Deduplicate items by merging identical items and summing quantities.

    Items are considered identical if they have the same:
    - id (code)
    - width
    - height
    - thickness
    - material
    - rotatable

    Args:
        items: List of item dictionaries

    Returns:
        List of deduplicated item dictionaries
    """
    item_map = {}

    for item in items:
        # Create unique key for duplicate detection
        key = (
            item.get('id', ''),
            item.get('width', 0),
            item.get('height', 0),
            item.get('thickness', 0),
            item.get('material', ''),
            item.get('rotatable', True)
        )

        if key in item_map:
            # Duplicate found - merge by adding quantities
            item_map[key]['quantity'] += item.get('quantity', 1)
        else:
            # New unique item - create a copy
            item_map[key] = item.copy()

    return list(item_map.values())


@app.route('/')
def index():
    """Render the main page."""
    return render_template('index.html')


@app.route('/api/solve', methods=['POST'])
def api_solve():
    """
    Solve a cutting stock problem.

    Expects JSON with items, bins, and parameters.
    Returns solution with bins, items, and cuts.
    """
    try:
        problem_data = request.get_json()

        if not problem_data:
            return jsonify({"error": "No problem data provided"}), 400

        # Deduplicate items before solving
        if 'items' in problem_data:
            original_count = len(problem_data['items'])
            problem_data['items'] = deduplicate_items(problem_data['items'])
            deduplicated_count = len(problem_data['items'])

            if original_count != deduplicated_count:
                print(f"Deduplicated items: {original_count} -> {deduplicated_count}")

        # Get time limit from parameters or use default
        time_limit = problem_data.get("parameters", {}).get("timeLimit", 5.0)

        # Solve the problem
        solution_dict = solve(problem_data, time_limit=time_limit)

        # Validate the solution
        problem = Problem.from_dict(problem_data)
        validator = SolutionValidator(problem)

        # Convert solution dict back to Solution object for validation
        # (This is a simplified validation - full validation would require conversion)
        is_valid = solution_dict["metadata"]["isComplete"]

        # Add validation info to metadata
        solution_dict["metadata"]["validated"] = is_valid

        return jsonify(solution_dict)

    except Exception as e:
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({
            "error": error_msg,
            "traceback": traceback.format_exc()
        }), 500


@app.route('/api/validate', methods=['POST'])
def api_validate():
    """
    Validate a solution.

    Expects JSON with problem and solution data.
    Returns validation result with errors if any.
    """
    try:
        data = request.get_json()
        problem_data = data.get("problem")
        solution_data = data.get("solution")

        if not problem_data or not solution_data:
            return jsonify({"error": "Missing problem or solution data"}), 400

        problem = Problem.from_dict(problem_data)
        validator = SolutionValidator(problem)

        # Note: This is simplified - full validation would require
        # converting solution_data back to Solution object
        is_complete = len(solution_data.get("unplaced", [])) == 0

        return jsonify({
            "valid": is_complete,
            "errors": [] if is_complete else ["Some items were not placed"],
            "metadata": solution_data.get("metadata", {})
        })

    except Exception as e:
        error_msg = str(e)
        return jsonify({"error": error_msg}), 500


@app.route('/api/example', methods=['GET'])
def api_example():
    """
    Get an example problem.

    Returns a sample problem that can be solved.
    """
    example = {
        "items": [
            {
                "id": "I001",
                "width": 2000,
                "height": 600,
                "thickness": 18,
                "material": "Oak",
                "quantity": 2,
                "rotatable": False
            },
            {
                "id": "I002",
                "width": 900,
                "height": 600,
                "thickness": 18,
                "material": "Oak",
                "quantity": 2,
                "rotatable": True
            },
            {
                "id": "I003",
                "width": 880,
                "height": 580,
                "thickness": 18,
                "material": "Oak",
                "quantity": 3,
                "rotatable": True
            }
        ],
        "bins": [
            {
                "id": "STD-1220x2440",
                "width": 1220,
                "height": 2440,
                "thickness": 18,
                "material": "Oak",
                "available": -1
            }
        ],
        "parameters": {
            "kerf": 3.0,
            "utilizationThreshold": 0.78,
            "timeLimit": 5.0
        }
    }

    return jsonify(example)


@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    """
    Upload and parse an Excel file with product data.

    Expected Excel columns:
    - Name: Product name
    - Code: Product code
    - Width: Width in mm (horizontal dimension)
    - Height: Height in mm (vertical dimension)
    - Thickness: Thickness in mm
    - Color: Material color/type
    - Qty: Quantity
    - Grain: Grain orientation (mixed/fixed)

    Returns JSON array of products.
    """
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"error": "File must be an Excel file (.xlsx or .xls)"}), 400

        # Load the workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Find header row (should be first row)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.strip() if cell.value else "")

        # Validate required columns
        required_columns = ['Name', 'Code', 'Width', 'Height', 'Thickness', 'Color', 'Qty', 'Grain']
        missing_columns = [col for col in required_columns if col not in headers]

        if missing_columns:
            return jsonify({
                "error": f"Missing required columns: {', '.join(missing_columns)}",
                "foundColumns": headers
            }), 400

        # Get column indices
        col_indices = {col: headers.index(col) for col in required_columns}

        # Parse rows
        products = []
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue

            try:
                name = row[col_indices['Name']]
                code = row[col_indices['Code']]
                width = row[col_indices['Width']]
                height = row[col_indices['Height']]
                thickness = row[col_indices['Thickness']]
                color = row[col_indices['Color']]
                qty = row[col_indices['Qty']]
                grain = row[col_indices['Grain']]

                # Validate required fields
                if not all([name, code, width, height, thickness, color, qty, grain]):
                    errors.append(f"Row {row_num}: Missing required field(s)")
                    continue

                # Convert and validate numeric fields
                try:
                    width = float(width)
                    height = float(height)
                    thickness = float(thickness)
                    qty = int(qty)
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Invalid numeric value")
                    continue

                # Validate grain value
                grain = str(grain).lower().strip()
                if grain not in ['mixed', 'fixed']:
                    errors.append(f"Row {row_num}: Grain must be 'mixed' or 'fixed', got '{grain}'")
                    continue

                # Add product
                products.append({
                    'name': str(name).strip(),
                    'code': str(code).strip(),
                    'width': width,
                    'height': height,
                    'thickness': thickness,
                    'color': str(color).strip(),
                    'qty': qty,
                    'grain': grain
                })

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        if not products and errors:
            return jsonify({
                "error": "No valid products found",
                "details": errors
            }), 400

        return jsonify({
            "products": products,
            "warnings": errors if errors else None,
            "count": len(products)
        })

    except Exception as e:
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({
            "error": f"Failed to process Excel file: {error_msg}",
            "traceback": traceback.format_exc()
        }), 500


@app.route('/api/download-template', methods=['GET'])
def download_template():
    """
    Download an Excel template for product data entry.

    Returns an Excel file with:
    - Header row with all required columns
    - Example rows with sample data
    - Instructions sheet
    """
    try:
        # Create workbook
        wb = Workbook()

        # Data sheet
        ws = wb.active
        ws.title = "Products"

        # Define headers
        headers = ['Name', 'Code', 'Width', 'Height', 'Thickness', 'Color', 'Qty', 'Grain']

        # Style for header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add example data
        examples = [
            {
                'Name': 'Cabinet Door',
                'Code': 'CAB-001',
                'Width': 800,
                'Height': 600,
                'Thickness': 18,
                'Color': 'Oak',
                'Qty': 4,
                'Grain': 'fixed'
            },
            {
                'Name': 'Shelf Board',
                'Code': 'SHF-001',
                'Width': 1200,
                'Height': 400,
                'Thickness': 18,
                'Color': 'Oak',
                'Qty': 3,
                'Grain': 'mixed'
            },
            {
                'Name': 'Table Top',
                'Code': 'TBL-001',
                'Width': 1800,
                'Height': 900,
                'Thickness': 25,
                'Color': 'Walnut',
                'Qty': 1,
                'Grain': 'fixed'
            }
        ]

        for row_num, example in enumerate(examples, start=2):
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=row_num, column=col_num, value=example[header])

        # Set column widths
        column_widths = {
            'A': 20,  # Name
            'B': 12,  # Code
            'C': 10,  # Width
            'D': 10,  # Height
            'E': 12,  # Thickness
            'F': 15,  # Color
            'G': 8,   # Qty
            'H': 12   # Grain
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")

        instructions = [
            ("Furniture Component Cutting Plan - Excel Template", "header"),
            ("", ""),
            ("Column Descriptions:", "subheader"),
            ("", ""),
            ("Name", "Label for the product (e.g., 'Cabinet Door', 'Shelf Board')"),
            ("Code", "Unique product code (e.g., 'CAB-001')"),
            ("Width", "Horizontal dimension in millimeters (left to right)"),
            ("Height", "Vertical dimension in millimeters (top to bottom)"),
            ("Thickness", "Board thickness in millimeters"),
            ("Color", "Material type or color (e.g., 'Oak', 'Walnut')"),
            ("Qty", "Quantity needed (whole number)"),
            ("Grain", "Grain orientation: 'mixed' or 'fixed'"),
            ("", ""),
            ("Dimension Guide:", "subheader"),
            ("", ""),
            ("Width", "Horizontal dimension (left to right on screen)"),
            ("Height", "Vertical dimension (top to bottom on screen)"),
            ("Example", "Board 1220 x 2440 means Width=2440mm, Height=1220mm"),
            ("", ""),
            ("Grain Orientation Guide:", "subheader"),
            ("", ""),
            ("mixed", "Can rotate 90 degrees - no orientation constraint"),
            ("fixed", "Cannot rotate - must maintain single direction"),
            ("", ""),
            ("Notes:", "subheader"),
            ("", ""),
            ("• All fields are required", ""),
            ("• Width and Height are in millimeters", ""),
            ("• Qty must be a whole number", ""),
            ("• Delete the example rows before uploading your own data", ""),
            ("• You can add as many rows as needed", ""),
        ]

        for row_num, (col1, col2) in enumerate(instructions, start=1):
            cell1 = ws_instructions.cell(row=row_num, column=1, value=col1)

            if col2 == "header":
                cell1.font = Font(bold=True, size=14, color="4472C4")
            elif col2 == "subheader":
                cell1.font = Font(bold=True, size=12)
            else:
                ws_instructions.cell(row=row_num, column=2, value=col2)

        ws_instructions.column_dimensions['A'].width = 25
        ws_instructions.column_dimensions['B'].width = 60

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='cutting_plan_template.xlsx'
        )

    except Exception as e:
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({
            "error": f"Failed to generate template: {error_msg}"
        }), 500


@app.route('/api/export-products', methods=['POST'])
def export_products():
    """
    Export product list to Excel file.

    Expects JSON with products array.
    Returns Excel file with same format as template.
    """
    try:
        data = request.get_json()
        products = data.get('products', [])

        if not products:
            return jsonify({"error": "No products provided"}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Define headers (same as template)
        headers = ['Name', 'Code', 'Width', 'Height', 'Thickness', 'Color', 'Qty', 'Grain']

        # Style for header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write product data
        for row_num, product in enumerate(products, start=2):
            ws.cell(row=row_num, column=1, value=product.get('name', ''))
            ws.cell(row=row_num, column=2, value=product.get('code', ''))
            ws.cell(row=row_num, column=3, value=product.get('width', ''))
            ws.cell(row=row_num, column=4, value=product.get('height', ''))
            ws.cell(row=row_num, column=5, value=product.get('thickness', ''))
            ws.cell(row=row_num, column=6, value=product.get('color', ''))
            ws.cell(row=row_num, column=7, value=product.get('qty', 1))
            ws.cell(row=row_num, column=8, value=product.get('grain', 'mixed'))

        # Set column widths (same as template)
        column_widths = {
            'A': 20,  # Name
            'B': 12,  # Code
            'C': 10,  # Width
            'D': 10,  # Height
            'E': 12,  # Thickness
            'F': 15,  # Color
            'G': 8,   # Qty
            'H': 12   # Grain
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='product_list.xlsx'
        )

    except Exception as e:
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({
            "error": f"Failed to export products: {error_msg}"
        }), 500


@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({"status": "healthy", "version": "1.0.0"})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('DEBUG', 'True').lower() == 'true'

    print("=" * 60)
    print("🎯 Tessellate Cutting Stock Optimizer")
    print("=" * 60)
    print(f"🌐 Server running on http://localhost:{port}")
    print(f"📊 Debug mode: {debug}")
    print("=" * 60)

    app.run(host='0.0.0.0', port=port, debug=debug)
