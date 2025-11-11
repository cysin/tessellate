#!/usr/bin/env python3
"""
Test Manual Pattern algorithm derived from manual1.jpg analysis.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Problem, Item, Bin
from tessellate.algorithms.manual_pattern_packer import ManualPatternPacker
from tessellate.algorithms.skyline import SkylinePacker


def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            normalized_headers[column_mapping[header]] = idx

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        grain_mapping = {'mixed': True, 'fixed': False, '可旋转': True, '不可旋转': False}

        item = Item(
            id=row[normalized_headers['Code']],
            width=int(float(row[normalized_headers['Width']])),
            height=int(float(row[normalized_headers['Height']])),
            thickness=int(float(row[normalized_headers['Thickness']])),
            material=row[normalized_headers['Color']],
            quantity=int(float(row[normalized_headers['Qty']])),
            rotatable=grain_mapping.get(row[normalized_headers['Grain']], True)
        )
        items.append(item)

    return items


def verify_solution(solution, expected_items):
    """Verify solution has all expected items."""

    # Count items in solution
    placed_items = {}
    for bin_packing in solution.bins:
        for placed_item in bin_packing.items:
            item_id = placed_item.item.id
            placed_items[item_id] = placed_items.get(item_id, 0) + 1

    # Expected items
    expected_counts = {}
    for item in expected_items:
        expected_counts[item.id] = item.quantity

    total_expected = sum(expected_counts.values())
    total_placed = sum(placed_items.values())

    all_correct = True
    for item_id, expected_qty in expected_counts.items():
        placed_qty = placed_items.get(item_id, 0)
        if placed_qty != expected_qty:
            all_correct = False

    return all_correct and total_placed == total_expected


def main():
    excel_path = Path("test_data/bench/manual1.xlsx")

    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        return

    print("=" * 70)
    print("Testing Manual Pattern Algorithm (derived from manual1.jpg)")
    print("=" * 70)
    print()

    items = load_excel_manual1(excel_path)

    print(f"Loaded {len(items)} item types")
    print(f"Total items: {sum(item.quantity for item in items)}")
    print()

    # Create bins
    thicknesses = list(set(item.thickness for item in items))
    materials = list(set(item.material for item in items))

    bins = []
    for thickness in thicknesses:
        for material in materials:
            bins.append(
                Bin(
                    id=f"Board-2440x1220-{thickness}mm-{material}",
                    width=2440,
                    height=1220,
                    thickness=thickness,
                    material=material
                )
            )

    problem = Problem(items=items, bins=bins, kerf=3.0)

    # Test algorithms
    print("=" * 70)
    print("TESTING")
    print("=" * 70)
    print()

    algorithms = [
        ("Manual Pattern", ManualPatternPacker(time_limit=5.0)),
        ("Skyline (baseline)", SkylinePacker(time_limit=5.0, use_min_waste=True)),
    ]

    results = []

    for name, algorithm in algorithms:
        print(f"Testing: {name}")
        print("-" * 70)

        solution = algorithm.solve(problem)

        num_bins = solution.num_bins()
        utilization = solution.total_utilization()
        items_placed = sum(len(bp.items) for bp in solution.bins)
        items_unplaced = sum(qty for _, qty in solution.unplaced)
        is_valid = verify_solution(solution, items)

        results.append({
            'name': name,
            'bins': num_bins,
            'utilization': utilization,
            'placed': items_placed,
            'unplaced': items_unplaced,
            'valid': is_valid
        })

        print(f"  Result: {num_bins} bins @ {utilization:.2%}")
        print(f"  Items: {items_placed} placed, {items_unplaced} unplaced")
        print(f"  Valid: {'✅ YES' if is_valid else '❌ NO'}")
        print()

    # Summary
    print()
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print()
    print(f"{'Algorithm':<25} {'Bins':>6} {'Util%':>7} {'Valid':>8}")
    print("-" * 70)

    for result in results:
        valid_str = '✅' if result['valid'] else '❌'
        print(f"{result['name']:<25} {result['bins']:>6} {result['utilization']:>6.2%} {valid_str:>8}")

    print()
    print("Target: 10 boards (manual solution from manual1.jpg)")
    print()

    # Find best result
    valid_results = [r for r in results if r['valid']]
    if valid_results:
        best = min(valid_results, key=lambda x: (x['bins'], -x['utilization']))
        print(f"Best result: {best['name']}")
        print(f"  {best['bins']} boards @ {best['utilization']:.2%} utilization")

        if best['bins'] == 10:
            print()
            print("🎉 🎉 🎉 SUCCESS! 🎉 🎉 🎉")
            print("MATCHED THE MANUAL SOLUTION TARGET!")
            print("The Manual Pattern algorithm derived from manual1.jpg")
            print("successfully replicates the human optimization!")
        elif best['bins'] == 11:
            print()
            print("Result: 11 boards (+1 vs manual target)")
            print("Very close to manual solution!")
        else:
            print()
            print(f"Result: {best['bins']} boards")


if __name__ == "__main__":
    main()
