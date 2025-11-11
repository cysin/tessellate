#!/usr/bin/env python3
"""
Load manual1.xlsx and test simplified ILP packer.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Add tessellate to path
sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Problem, Item, Bin
from tessellate.algorithms.ilp_simplified_packer import SimplifiedILPPacker


def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Find header row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value.strip() if cell.value else "")

    # Map column names
    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            english_name = column_mapping[header]
            normalized_headers[english_name] = idx

    # Parse rows
    items = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            name = row[normalized_headers['Name']]
            code = row[normalized_headers['Code']]
            width = int(float(row[normalized_headers['Width']]))
            height = int(float(row[normalized_headers['Height']]))
            thickness = int(float(row[normalized_headers['Thickness']]))
            color = row[normalized_headers['Color']]
            qty = int(float(row[normalized_headers['Qty']]))
            grain = row[normalized_headers['Grain']]

            # Convert grain to rotatable
            grain_mapping = {
                'mixed': True,
                'fixed': False,
                '可旋转': True,
                '不可旋转': False
            }
            rotatable = grain_mapping.get(grain, True)

            item = Item(
                id=code,
                width=width,
                height=height,
                thickness=thickness,
                material=color,
                quantity=qty,
                rotatable=rotatable
            )
            items.append(item)

        except Exception as e:
            print(f"Error parsing row {row_num}: {e}")

    return items


def main():
    excel_path = Path("test_data/bench/manual1.xlsx")

    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        return

    print("=" * 70)
    print("Loading manual1.xlsx...")
    print("=" * 70)

    items = load_excel_manual1(excel_path)

    print(f"Loaded {len(items)} item types")
    print(f"Total items: {sum(item.quantity for item in items)}")
    print()

    # Create bins (standard 2440x1220 boards)
    # Get unique thicknesses from items
    thicknesses = list(set(item.thickness for item in items))
    materials = list(set(item.material for item in items))

    print(f"Item thicknesses: {thicknesses}")
    print(f"Item materials: {materials}")
    print()

    # Create bins for each thickness/material combination
    bins = []
    for thickness in thicknesses:
        for material in materials:
            bins.append(
                Bin(
                    id=f"Board-2440x1220-{thickness}mm-{material}",
                    width=2440,
                    height=1220,
                    thickness=thickness,
                    material=material
                )
            )

    # Create problem
    problem = Problem(
        items=items,
        bins=bins,
        kerf=3.0
    )

    print("=" * 70)
    print("Testing Simplified ILP Packer")
    print("=" * 70)
    print(f"Target: 10 boards")
    print()

    # Try simplified ILP targeting 10 bins
    solver = SimplifiedILPPacker(
        time_limit=300.0,  # 5 minutes
        target_bins=10,
        mip_gap=0.05
    )

    solution = solver.solve(problem)

    print()
    print("=" * 70)
    print("RESULTS")
    print("=" * 70)
    print(f"Bins: {solution.num_bins()}")
    print(f"Utilization: {solution.total_utilization():.2%}")
    print(f"Items placed: {sum(len(bp.items) for bp in solution.bins)}")
    print(f"Items unplaced: {sum(qty for _, qty in solution.unplaced)}")

    if solution.num_bins() == 10 and len(solution.unplaced) == 0:
        print()
        print("🎉 SUCCESS: Achieved 10 boards target!")
    elif solution.num_bins() > 0:
        print()
        print(f"Result: {solution.num_bins()} boards (target was 10)")
    else:
        print()
        print("❌ ILP solver could not find solution")


if __name__ == "__main__":
    main()
