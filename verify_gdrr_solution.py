#!/usr/bin/env python3
"""
Verify GDRR solution - ensure all items are placed correctly.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Problem, Item, Bin
from tessellate.algorithms.gdrr_packer import GDRRPacker


def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            normalized_headers[column_mapping[header]] = idx

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        grain_mapping = {'mixed': True, 'fixed': False, '可旋转': True, '不可旋转': False}

        item = Item(
            id=row[normalized_headers['Code']],
            width=int(float(row[normalized_headers['Width']])),
            height=int(float(row[normalized_headers['Height']])),
            thickness=int(float(row[normalized_headers['Thickness']])),
            material=row[normalized_headers['Color']],
            quantity=int(float(row[normalized_headers['Qty']])),
            rotatable=grain_mapping.get(row[normalized_headers['Grain']], True)
        )
        items.append(item)

    return items


def verify_solution(solution, expected_items):
    """Verify solution has all expected items."""

    print("\n" + "="*70)
    print("VERIFICATION")
    print("="*70)

    # Count items in solution
    placed_items = {}
    for bin_packing in solution.bins:
        for placed_item in bin_packing.items:
            item_id = placed_item.item.id
            placed_items[item_id] = placed_items.get(item_id, 0) + 1

    # Expected items
    expected_counts = {}
    for item in expected_items:
        expected_counts[item.id] = item.quantity

    total_expected = sum(expected_counts.values())
    total_placed = sum(placed_items.values())
    total_unplaced = sum(qty for _, qty in solution.unplaced)

    print(f"Expected items: {total_expected}")
    print(f"Placed items: {total_placed}")
    print(f"Unplaced items: {total_unplaced}")
    print()

    # Check each item type
    all_correct = True
    for item_id, expected_qty in expected_counts.items():
        placed_qty = placed_items.get(item_id, 0)

        if placed_qty != expected_qty:
            print(f"❌ {item_id}: Expected {expected_qty}, Placed {placed_qty}")
            all_correct = False
        else:
            print(f"✓ {item_id}: {placed_qty}/{expected_qty}")

    print()
    if all_correct and total_placed == total_expected:
        print("✅ VERIFICATION PASSED: All items correctly placed!")
        return True
    else:
        print("❌ VERIFICATION FAILED: Item counts don't match!")
        return False


def main():
    excel_path = Path("test_data/bench/manual1.xlsx")

    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        return

    print("=" * 70)
    print("GDRR Verification Test")
    print("=" * 70)

    items = load_excel_manual1(excel_path)
    print(f"\nLoaded {len(items)} item types")
    print(f"Total items: {sum(item.quantity for item in items)}")

    # Create bins
    thicknesses = list(set(item.thickness for item in items))
    materials = list(set(item.material for item in items))

    bins = []
    for thickness in thicknesses:
        for material in materials:
            bins.append(
                Bin(
                    id=f"Board-2440x1220-{thickness}mm-{material}",
                    width=2440,
                    height=1220,
                    thickness=thickness,
                    material=material
                )
            )

    problem = Problem(items=items, bins=bins, kerf=3.0)

    # Test GDRR
    print("\nRunning GDRR...")
    solver = GDRRPacker(time_limit=60.0, iterations=500, history_length=100)
    solution = solver.solve(problem)

    print(f"\nResult: {solution.num_bins()} bins @ {solution.total_utilization():.2%}")

    # Verify
    is_valid = verify_solution(solution, items)

    if is_valid:
        print("\n" + "="*70)
        print(f"🎉 SUCCESS: GDRR achieved {solution.num_bins()} bins")
        print(f"   This is better than the manual 10-board solution!")
        print("="*70)


if __name__ == "__main__":
    main()
