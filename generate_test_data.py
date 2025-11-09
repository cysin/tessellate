#!/usr/bin/env python3
"""
Generate test Excel files from components_data.csv.

This script simulates realistic wardrobe orders by selecting components
from the CSV file and generating Excel files with random quantities.

Grain orientation is determined by color code:
- HS00, HS99, HS98, HS97 -> 'mixed' (can rotate 90 degrees)
- HS01, HS02, HS03 -> 'fixed' (cannot rotate)

Usage:
    python generate_test_data.py <number_of_files>

Example:
    python generate_test_data.py 5
"""

import sys
import csv
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class Component:
    """Represents a wardrobe component."""

    def __init__(self, name, code):
        self.name = name
        self.code = code
        self.parse_code()

    def parse_code(self):
        """Parse the component code to extract color, width, height, thickness."""
        # Format: PREFIX-COLOR-WIDTH-HEIGHT-THICK
        # Example: CB(L)-HS00-2434-574-16
        parts = self.code.split('-')

        if len(parts) >= 5:
            # Last 4 parts are COLOR-WIDTH-HEIGHT-THICK
            self.prefix = '-'.join(parts[:-4])
            self.color = parts[-4]
            self.width = int(parts[-3])
            self.height = int(parts[-2])
            self.thick = int(parts[-1])
        else:
            # Fallback if format doesn't match
            self.prefix = parts[0] if parts else ""
            self.color = "HS00"
            self.width = 800
            self.height = 600
            self.thick = 18

    def __repr__(self):
        return f"Component({self.name}, {self.code})"


def load_components(csv_file):
    """Load components from CSV file."""
    components = []

    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 2:
                name = row[0].strip()
                code = row[1].strip()
                components.append(Component(name, code))

    return components


def categorize_components(components):
    """Categorize components by type for realistic selection."""
    categories = {
        '侧板': [],      # Side panels
        '顶底板': [],    # Top/bottom panels
        '层隔板': [],    # Shelves
        '背板': [],      # Back panels
        '门板': [],      # Doors
        '抽屉': [],      # Drawers
        '其他': []       # Others
    }

    for comp in components:
        categorized = False
        for key in categories.keys():
            if key in comp.name:
                categories[key].append(comp)
                categorized = True
                break

        if not categorized:
            categories['其他'].append(comp)

    return categories


def generate_realistic_order(categories):
    """Generate a realistic wardrobe order with appropriate quantities."""
    order = []

    # 1. Side panels - typically 2 (left and right)
    if categories['侧板']:
        side_panels = random.sample(categories['侧板'], min(2, len(categories['侧板'])))
        for panel in side_panels:
            order.append((panel, 1))  # Usually 1 of each side

    # 2. Top/bottom panels - typically 2-4 pieces (single or double door)
    if categories['顶底板']:
        top_bottom = random.sample(categories['顶底板'], random.randint(2, min(4, len(categories['顶底板']))))
        for panel in top_bottom:
            order.append((panel, random.randint(1, 2)))

    # 3. Shelves - typically 3-8 pieces
    if categories['层隔板']:
        shelves = random.sample(categories['层隔板'], random.randint(3, min(8, len(categories['层隔板']))))
        for shelf in shelves:
            order.append((shelf, random.randint(1, 3)))

    # 4. Back panels - typically 1-2 pieces
    if categories['背板']:
        back_panels = random.sample(categories['背板'], random.randint(1, min(2, len(categories['背板']))))
        for panel in back_panels:
            order.append((panel, 1))

    # 5. Doors - typically 1-4 pieces
    if categories['门板']:
        doors = random.sample(categories['门板'], random.randint(1, min(4, len(categories['门板']))))
        for door in doors:
            order.append((door, random.randint(1, 2)))

    # 6. Drawers - optional, 0-4 sets
    if categories['抽屉'] and random.random() > 0.3:
        drawers = random.sample(categories['抽屉'], random.randint(1, min(4, len(categories['抽屉']))))
        for drawer in drawers:
            order.append((drawer, random.randint(1, 3)))

    # 7. Other components - randomly add 2-5 items
    if categories['其他']:
        others = random.sample(categories['其他'], random.randint(2, min(5, len(categories['其他']))))
        for other in others:
            order.append((other, random.randint(1, 2)))

    return order


def create_excel_file(order, filename):
    """Create an Excel file from the order."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Define headers
    headers = ['Name', 'Code', 'Width', 'Height', 'Thickness', 'Color', 'Qty', 'Grain']

    # Style for header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, (component, qty) in enumerate(order, start=2):
        # Determine grain based on color code
        # HS00, HS99, HS98, HS97 -> mixed (can rotate)
        # HS01, HS02, HS03 -> fixed (cannot rotate)
        if component.color in ['HS00', 'HS99', 'HS98', 'HS97']:
            grain = 'mixed'
        else:  # HS01, HS02, HS03 or any other color
            grain = 'fixed'

        ws.cell(row=row_num, column=1, value=component.name)
        ws.cell(row=row_num, column=2, value=component.code)
        ws.cell(row=row_num, column=3, value=component.width)
        ws.cell(row=row_num, column=4, value=component.height)
        ws.cell(row=row_num, column=5, value=component.thick)
        ws.cell(row=row_num, column=6, value=component.color)
        ws.cell(row=row_num, column=7, value=qty)
        ws.cell(row=row_num, column=8, value=grain)

    # Set column widths
    column_widths = {
        'A': 30,  # Name
        'B': 25,  # Code
        'C': 10,  # Width
        'D': 10,  # Height
        'E': 12,  # Thickness
        'F': 12,  # Color
        'G': 8,   # Qty
        'H': 12   # Grain
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save
    wb.save(filename)
    print(f"✓ Generated: {filename}")


def main():
    """Main function."""
    if len(sys.argv) != 2:
        print("Usage: python generate_test_data.py <number_of_files>")
        print("Example: python generate_test_data.py 5")
        sys.exit(1)

    try:
        num_files = int(sys.argv[1])
        if num_files <= 0:
            raise ValueError("Number must be positive")
    except ValueError as e:
        print(f"Error: Invalid number - {e}")
        sys.exit(1)

    # Load components
    csv_file = Path(__file__).parent / "components_data.csv"
    if not csv_file.exists():
        print(f"Error: {csv_file} not found")
        sys.exit(1)

    print(f"Loading components from {csv_file}...")
    components = load_components(csv_file)
    print(f"Loaded {len(components)} components")

    # Categorize components
    categories = categorize_components(components)
    print(f"\nComponent categories:")
    for cat, comps in categories.items():
        if comps:
            print(f"  {cat}: {len(comps)} items")

    # Create output directory
    output_dir = Path(__file__).parent / "test_data"
    output_dir.mkdir(exist_ok=True)

    # Generate test files
    print(f"\nGenerating {num_files} test files...")
    for i in range(1, num_files + 1):
        # Generate realistic order
        order = generate_realistic_order(categories)

        # Create filename
        filename = output_dir / f"wardrobe_order_{i:03d}.xlsx"

        # Create Excel file
        create_excel_file(order, filename)

    print(f"\n✅ Successfully generated {num_files} test files in {output_dir}/")
    print(f"\nSummary:")
    print(f"  Total components available: {len(components)}")
    print(f"  Files generated: {num_files}")
    print(f"  Output directory: {output_dir}")


if __name__ == "__main__":
    main()
