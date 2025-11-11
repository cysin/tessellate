#!/usr/bin/env python3
"""
Test ILP Decomposition packer on manual1.xlsx.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from tessellate.core.models import Problem, Item, Bin
from tessellate.algorithms.ilp_decomposition_packer import ILPDecompositionPacker


def load_excel_manual1(excel_path):
    """Load the manual1.xlsx file."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

    column_mapping = {
        'Name': 'Name', '名称': 'Name',
        'Code': 'Code', '编码': 'Code',
        'Width': 'Width', '宽度': 'Width',
        'Height': 'Height', '高度': 'Height',
        'Thickness': 'Thickness', '厚度': 'Thickness',
        'Color': 'Color', '颜色': 'Color',
        'Qty': 'Qty', '数量': 'Qty',
        'Grain': 'Grain', '纹理': 'Grain'
    }

    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header in column_mapping:
            normalized_headers[column_mapping[header]] = idx

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        grain_mapping = {'mixed': True, 'fixed': False, '可旋转': True, '不可旋转': False}

        item = Item(
            id=row[normalized_headers['Code']],
            width=int(float(row[normalized_headers['Width']])),
            height=int(float(row[normalized_headers['Height']])),
            thickness=int(float(row[normalized_headers['Thickness']])),
            material=row[normalized_headers['Color']],
            quantity=int(float(row[normalized_headers['Qty']])),
            rotatable=grain_mapping.get(row[normalized_headers['Grain']], True)
        )
        items.append(item)

    return items


def main():
    excel_path = Path("test_data/bench/manual1.xlsx")

    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        return

    print("=" * 70)
    print("ILP Decomposition Approach")
    print("=" * 70)

    items = load_excel_manual1(excel_path)

    print(f"Loaded {len(items)} item types")
    print(f"Total items: {sum(item.quantity for item in items)}")
    print()

    # Create bins matching item specifications
    thicknesses = list(set(item.thickness for item in items))
    materials = list(set(item.material for item in items))

    bins = []
    for thickness in thicknesses:
        for material in materials:
            bins.append(
                Bin(
                    id=f"Board-2440x1220-{thickness}mm-{material}",
                    width=2440,
                    height=1220,
                    thickness=thickness,
                    material=material
                )
            )

    problem = Problem(items=items, bins=bins, kerf=3.0)

    # Solve with decomposition
    solver = ILPDecompositionPacker(
        time_limit=300.0,
        group_size=20,  # 20 items per subproblem (manageable for ILP)
        mip_gap=0.05
    )

    solution = solver.solve(problem)

    print()
    print("=" * 70)
    print("RESULTS")
    print("=" * 70)
    print(f"Bins: {solution.num_bins()}")
    print(f"Utilization: {solution.total_utilization():.2%}")
    print(f"Items placed: {sum(len(bp.items) for bp in solution.bins)}")
    print(f"Items unplaced: {sum(qty for _, qty in solution.unplaced)}")

    if solution.num_bins() <= 10 and len(solution.unplaced) == 0:
        print()
        print(f"🎉 Excellent result: {solution.num_bins()} boards!")
        if solution.num_bins() == 10:
            print("   MATCHED manual solution target!")
    elif solution.num_bins() > 0:
        print()
        print(f"Result: {solution.num_bins()} boards (manual target: 10)")


if __name__ == "__main__":
    main()
